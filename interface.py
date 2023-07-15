from tkinter import *
from tkinter import messagebox, filedialog
import openpyxl

# TODO: Save the new data to python list
# TODO: Open excel file and append and save the new data
# TODO: Validate closing and saving of data to excel sheet

raw_data = []


def next_data_btn():
    # Insert from validation
    simsval = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    new = ["", patient_name.get(), temp.get()] + simsval
    # write and append it into a data set
    if len(new[1]) == 0 or len(new[2]) == 0:
        messagebox.showinfo(title="Empty field", message="Please ensure all the fields arent empty")
    else:
        raw_data.append(new)
        patient_name.delete(0, END)
        patient_name.focus()
        temp.delete(0, END)


def save_sheet_btn():
    selected_sheet = ""

    def select_sheet_menu():
        selected_sheet = s1.get(ACTIVE)
        pick_sheet.destroy()

    # Input validation to check if the last data was saved

    last = raw_data[-1]
    if last[1] == patient_name.get() or patient_name.get() == "":
        # Opens the Excel file
        filename = filedialog.askopenfilename()
        filename.replace("/", "\\")
        excel_text = openpyxl.load_workbook(filename=filename)
        all_sheets = excel_text.sheetnames
        excel_text.close()

        # Select sheet window
        pick_sheet = Toplevel(window)
        pick_sheet.title("Select the sheet")
        pick_sheet.config(padx=50, pady=50, bg="white")

        # Scrollbar
        pick_sheet.yScroll = Scrollbar(pick_sheet, orient=VERTICAL)
        pick_sheet.yScroll.grid(row=0, column=1)

        # Select sheet to write file into
        s1 = Listbox(pick_sheet, bg="white", selectmode=SINGLE, yscrollcommand=pick_sheet.yScroll.set)
        s1.grid(row=0, column=0)
        pick_sheet.yScroll['command'] = s1.yview
        for i in range(len(all_sheets)):
            s1.insert(i + 1, all_sheets[i])

        select_sheet_btn = Button(pick_sheet, text="Select Sheet", command=select_sheet_menu)
        select_sheet_btn.grid(row=1, column=0)
        pick_sheet.mainloop()

        messagebox.showinfo(title="successful", message="save to Excel file successful")

        # Reads and displays the sheets in file
        """
        final_edit = openpyxl.load_workbook(filename=filename)
        editable_sheet = final_edit[selected_sheet]

        # Writes into sheet as a loop
        for i in raw_data:
            editable_sheet.append(i)

        # Saves file
        final_edit.save(filename)
        # Show confirmation message
        """

        # Close window

    else:
        next_data_btn()


# UI SETUP

window = Tk()
window.title("Paelon data entry")
window.config(padx=50, pady=50, bg="white")

# LABELS
name_label = Label(text="Patient Name:", bg="white")
name_label.grid(column=0, row=0)

temp_label = Label(text="Temperature:", bg="white")
temp_label.grid(column=0, row=1)

# BLANK SPACE
blank_label = Label(text="", bg="white")
blank_label.grid(column=0, row=2)

# DATA ENTRY

patient_name = Entry(width=27)
patient_name.grid(column=1, row=0, columnspan=2)
patient_name.focus()

temp = Entry(width=27)
temp.grid(column=1, row=1, columnspan=2)

# BUTTONS

next_data = Button(text="Next data", width=40, command=next_data_btn)
next_data.grid(column=1, row=3, columnspan=2)

save_data = Button(text="End", width=15, command=save_sheet_btn)
save_data.grid(column=0, row=3)

window.mainloop()
